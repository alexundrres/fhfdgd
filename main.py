from PyQt6 import QtCore, QtGui, QtWidgets
from PyQt6.QtWidgets import QApplication, QMainWindow, QDialog, QVBoxLayout, QLabel, QPushButton, QTableView, QMessageBox
from PyQt6.QtGui import QFont
from PyQt6.QtCore import QMetaObject, Qt, QRect
from PyQt6.QtSql import QSqlTableModel
import sys
from PyQt6.QtSql import QSqlDatabase, QSqlQuery
import openpyxl


class BD:
    def __init__(self):
        super(BD, self).__init__()
        self.create_tables()

    def create_tables(self):
        db = QSqlDatabase.addDatabase('QSQLITE')
        db.setDatabaseName("portal.db")
        if not db.open():
            print("Failed to open database")
            return False

        query = QSqlQuery()

        # Таблица "ученики"
        query.exec("""CREATE TABLE IF NOT EXISTS ученики (
                      id INTEGER PRIMARY KEY AUTOINCREMENT,
                      имя TEXT,
                      фамилия TEXT,
                      логин TEXT,
                      пароль TEXT DEFAULT no)""")

        # Таблица "преподаватели"
        query.exec("""CREATE TABLE IF NOT EXISTS преподаватели (
                      id INTEGER PRIMARY KEY AUTOINCREMENT,
                      имя TEXT,
                      фамилия TEXT,
                      логин TEXT,
                      пароль TEXT DEFAULT no)""")

        # Таблица "успеваемость"
        query.exec("""CREATE TABLE IF NOT EXISTS успеваемость (
                      id INTEGER PRIMARY KEY AUTOINCREMENT,
                      id_ученика INTEGER,
                      id_преподавателя INTEGER,
                      предмет TEXT,
                      оценка INTEGER,
                      FOREIGN KEY (id_ученика) REFERENCES ученики(id),
                      FOREIGN KEY (id_преподавателя) REFERENCES преподаватели(id))""")

        if not query.lastError().isValid():
            print("Tables created successfully")
        else:
            print("Error creating tables:", query.lastError().text())

        return True

    def insert_data(self, table, data):
        query = QSqlQuery()
        placeholders = ', '.join([':' + key for key in data.keys()])
        columns = ', '.join(data.keys())
        query.prepare(f"INSERT INTO {table} ({columns}) VALUES ({placeholders})")

        for key, value in data.items():
            query.bindValue(':' + key, value)

        success = query.exec()

        if not query.lastError().isValid():
            print("Data inserted successfully")
        else:
            print("Error inserting data:", query.lastError().text())

        return success

class Ui_avtorizacia(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(350, 400)
        MainWindow.setMinimumSize(QtCore.QSize(350, 400))
        MainWindow.setMaximumSize(QtCore.QSize(350, 400))
        MainWindow.setStyleSheet("")
        self.centralwidget = QtWidgets.QWidget(parent=MainWindow)
        self.centralwidget.setStyleSheet("background-color: rgb(255, 204, 153);")
        self.centralwidget.setObjectName("centralwidget")
        self.verticalLayout_2 = QtWidgets.QVBoxLayout(self.centralwidget)
        self.verticalLayout_2.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout_2.setObjectName("verticalLayout_2")
        self.frame = QtWidgets.QFrame(parent=self.centralwidget)
        self.frame.setMaximumSize(QtCore.QSize(500, 500))
        font = QtGui.QFont()
        font.setStyleStrategy(QtGui.QFont.StyleStrategy.PreferDefault)
        self.frame.setFont(font)
        self.frame.setLayoutDirection(QtCore.Qt.LayoutDirection.LeftToRight)
        self.frame.setAutoFillBackground(False)
        self.frame.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.frame.setFrameShape(QtWidgets.QFrame.Shape.StyledPanel)
        self.frame.setFrameShadow(QtWidgets.QFrame.Shadow.Raised)
        self.frame.setObjectName("frame")
        self.label = QtWidgets.QLabel(parent=self.frame)
        self.label.setGeometry(QtCore.QRect(110, 60, 121, 41))
        font = QtGui.QFont()
        font.setPointSize(14)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.pushButton = QtWidgets.QPushButton(parent=self.frame)
        self.pushButton.setGeometry(QtCore.QRect(80, 340, 190, 31))
        self.pushButton.setObjectName("pushButton")
        self.layoutWidget = QtWidgets.QWidget(parent=self.frame)
        self.layoutWidget.setGeometry(QtCore.QRect(100, 190, 141, 121))
        self.layoutWidget.setObjectName("layoutWidget")
        self.verticalLayout = QtWidgets.QVBoxLayout(self.layoutWidget)
        self.verticalLayout.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout.setObjectName("verticalLayout")
        self.lineEdit = QtWidgets.QLineEdit(parent=self.layoutWidget)
        self.lineEdit.setMinimumSize(QtCore.QSize(0, 30))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.lineEdit.setFont(font)
        self.lineEdit.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.lineEdit.setClearButtonEnabled(False)
        self.lineEdit.setObjectName("lineEdit")
        self.verticalLayout.addWidget(self.lineEdit)
        self.lineEdit_2 = QtWidgets.QLineEdit(parent=self.layoutWidget)
        self.lineEdit_2.setMinimumSize(QtCore.QSize(0, 30))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.lineEdit_2.setFont(font)
        self.lineEdit_2.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.lineEdit_2.setObjectName("lineEdit_2")
        self.verticalLayout.addWidget(self.lineEdit_2)
        self.comboBox = QtWidgets.QComboBox(parent=self.frame)
        self.comboBox.setGeometry(QtCore.QRect(110, 130, 121, 22))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.comboBox.setFont(font)
        self.comboBox.setContextMenuPolicy(QtCore.Qt.ContextMenuPolicy.NoContextMenu)
        self.comboBox.setLayoutDirection(QtCore.Qt.LayoutDirection.LeftToRight)
        self.comboBox.setInsertPolicy(QtWidgets.QComboBox.InsertPolicy.InsertAtBottom)
        self.comboBox.setObjectName("comboBox")
        self.comboBox.addItem("")
        self.comboBox.addItem("")
        self.verticalLayout_2.addWidget(self.frame)
        MainWindow.setCentralWidget(self.centralwidget)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

        self.pushButton.clicked.connect(self.login)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.label.setText(_translate("MainWindow", "Авторизация"))
        self.pushButton.setText(_translate("MainWindow", "Войти"))
        self.lineEdit.setPlaceholderText(_translate("MainWindow", "Логин"))
        self.lineEdit_2.setPlaceholderText(_translate("MainWindow", "Пароль"))
        self.comboBox.setItemText(0, _translate("MainWindow", "Ученик"))
        self.comboBox.setItemText(1, _translate("MainWindow", "Преподаватель"))

    def login(self):
        # Получаем введенные логин и пароль
        login = self.lineEdit.text()
        password = self.lineEdit_2.text()

        # Проверяем роль (ученик или преподаватель)
        role = self.comboBox.currentText()

        # Выполняем проверку в базе данных или другими способами
        # В данном примере просто сравниваем с фиксированными значениями
        if role == "Ученик" and self.check_credentials("ученики", login, password):
            self.show_student_window()
        elif role == "Преподаватель" and self.check_credentials("преподаватели", login, password):
            self.show_teacher_window()
        else:
            QMessageBox.warning(self.centralwidget, "Ошибка", "Неверные логин или пароль")

    def check_credentials(self, table, login, password):
        # Проверка учетных данных в базе данных
        query = QSqlQuery()
        query.prepare(f"SELECT * FROM {table} WHERE логин = :login AND пароль = :password")
        query.bindValue(":login", login)
        query.bindValue(":password", password)
        query.exec()
        return query.next()

    def show_student_window(self):
        self.MainWindow = QMainWindow()
        self.ui = Ui_ychenik()
        self.ui.setupUi(self.MainWindow)
        self.MainWindow.show()

    def show_teacher_window(self):
        self.MainWindow = QMainWindow()
        self.ui = Ui_prepodavatel()
        self.ui.setupUi(self.MainWindow)
        self.MainWindow.show()

class Ui_prepodavatel(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(800, 597)
        self.centralwidget = QtWidgets.QWidget(parent=MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.horizontalLayout = QtWidgets.QHBoxLayout(self.centralwidget)
        self.horizontalLayout.setContentsMargins(0, 0, 0, 0)
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.frame = QtWidgets.QFrame(parent=self.centralwidget)
        self.frame.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.frame.setFrameShape(QtWidgets.QFrame.Shape.StyledPanel)
        self.frame.setFrameShadow(QtWidgets.QFrame.Shadow.Raised)
        self.frame.setObjectName("frame")
        self.btn_exit = QtWidgets.QPushButton(parent=self.frame)
        self.btn_exit.setGeometry(QtCore.QRect(730, 0, 75, 23))
        self.btn_exit.setObjectName("btn_exit")
        self.btn_spravka = QtWidgets.QPushButton(parent=self.frame)
        self.btn_spravka.setGeometry(QtCore.QRect(660, 0, 75, 23))
        self.btn_spravka.setObjectName("btn_spravka")
        self.tableView_prep = QtWidgets.QTableView(parent=self.frame)
        self.tableView_prep.setGeometry(QtCore.QRect(60, 150, 671, 381))
        self.tableView_prep.setObjectName("tableView_prep")
        self.label = QtWidgets.QLabel(parent=self.frame)
        self.label.setGeometry(QtCore.QRect(60, 110, 91, 21))
        font = QtGui.QFont()
        font.setPointSize(14)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.btn_excel = QtWidgets.QPushButton(parent=self.frame)
        self.btn_excel.setGeometry(QtCore.QRect(60, 550, 75, 23))
        self.btn_excel.setObjectName("btn_excel")
        self.btn_delit = QtWidgets.QPushButton(parent=self.frame)
        self.btn_delit.setGeometry(QtCore.QRect(650, 120, 81, 31))
        self.btn_delit.setObjectName("btn_delit")
        self.btn_add = QtWidgets.QPushButton(parent=self.frame)
        self.btn_add.setGeometry(QtCore.QRect(570, 120, 81, 31))
        self.btn_add.setObjectName("btn_add")
        self.horizontalLayout.addWidget(self.frame)
        MainWindow.setCentralWidget(self.centralwidget)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def exit_button_clicked(self):
        self.frame.close()


    def spravka_button_clicked(self):
        QtWidgets.QMessageBox.information(self.frame, "Справка", "Это справочная информация.")


    def export_to_excel(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Пример: Запись данных из таблицы в Excel
        for row in range(self.tableView_prep.model().rowCount()):
            for column in range(self.tableView_prep.model().columnCount()):
                item = self.tableView_prep.model().item(row, column)
                sheet.cell(row=row + 1, column=column + 1, value=str(item.text()))


    def delete_record(self):
        selected_row = self.tableView_prep.selectionModel().currentIndex().row()
        self.tableView_prep.model().removeRow(selected_row)
        self.tableView_prep.model().submitAll()

    def add_record(self):
        row_position = self.tableView_prep.model().rowCount()
        self.tableView_prep.model().insertRow(row_position)
        self.tableView_prep.model().setData(self.tableView_prep.model().index(row_position, 0), "Новая запись")
        self.tableView_prep.model().submitAll()

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.btn_exit.setText(_translate("MainWindow", "Выход"))
        self.btn_spravka.setText(_translate("MainWindow", "Справка"))
        self.label.setText(_translate("MainWindow", "Отметки:"))
        self.btn_excel.setText(_translate("MainWindow", "Вывод в Excel"))
        self.btn_delit.setText(_translate("MainWindow", "Удалить"))
        self.btn_add.setText(_translate("MainWindow", "Добавить"))

class Ui_ychenik(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(800, 600)
        font = QtGui.QFont()
        font.setPointSize(10)
        MainWindow.setFont(font)
        self.centralwidget = QtWidgets.QWidget(parent=MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.horizontalLayout = QtWidgets.QHBoxLayout(self.centralwidget)
        self.horizontalLayout.setContentsMargins(0, 0, 0, 0)
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.frame = QtWidgets.QFrame(parent=self.centralwidget)
        self.frame.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.frame.setFrameShape(QtWidgets.QFrame.Shape.StyledPanel)
        self.frame.setFrameShadow(QtWidgets.QFrame.Shadow.Raised)
        self.frame.setObjectName("frame")
        self.btn_spravka = QtWidgets.QPushButton(parent=self.frame)
        self.btn_spravka.setGeometry(QtCore.QRect(644, 0, 71, 23))
        self.btn_spravka.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.btn_spravka.setObjectName("btn_spravka")
        self.btn_exit = QtWidgets.QPushButton(parent=self.frame)
        self.btn_exit.setGeometry(QtCore.QRect(720, 0, 75, 23))
        self.btn_exit.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.btn_exit.setObjectName("btn_exit")
        self.label = QtWidgets.QLabel(parent=self.frame)
        self.label.setGeometry(QtCore.QRect(60, 110, 91, 21))
        font = QtGui.QFont()
        font.setPointSize(14)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.tableView = QtWidgets.QTableView(parent=self.frame)
        self.tableView.setGeometry(QtCore.QRect(60, 150, 671, 381))
        self.tableView.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.tableView.setObjectName("tableView")
        self.btn_excel = QtWidgets.QPushButton(parent=self.frame)
        self.btn_excel.setGeometry(QtCore.QRect(60, 550, 81, 23))
        self.btn_excel.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.btn_excel.setObjectName("btn_excel")
        self.horizontalLayout.addWidget(self.frame)
        MainWindow.setCentralWidget(self.centralwidget)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def exit_button_clicked(self):
        self.frame.close()


    def spravka_button_clicked(self):
        QtWidgets.QMessageBox.information(self.frame, "Справка", "Это справочная информация.")


    def export_to_excel(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Пример: Запись данных из таблицы в Excel
        for row in range(self.tableView_prep.model().rowCount()):
            for column in range(self.tableView_prep.model().columnCount()):
                item = self.tableView_prep.model().item(row, column)
                sheet.cell(row=row + 1, column=column + 1, value=str(item.text()))

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.btn_spravka.setText(_translate("MainWindow", "Справка"))
        self.btn_exit.setText(_translate("MainWindow", "Выход"))
        self.label.setText(_translate("MainWindow", "Отметки:"))
        self.btn_excel.setText(_translate("MainWindow", "Вывод в Excel"))

class Main(QMainWindow):
    def __init__(self):
        super().__init__()

        self.avtorizacia = QMainWindow()
        self.ui_avtorizacia = Ui_avtorizacia()
        self.ui_avtorizacia.setupUi(self.avtorizacia)

        self.bd = BD()

        self.prepodavatel = QMainWindow()
        self.ui_prepodavatel = Ui_prepodavatel()
        self.ui_prepodavatel.setupUi(self.prepodavatel)

        self.ychenik = QMainWindow()
        self.ui_ychenik = Ui_ychenik()
        self.ui_ychenik.setupUi(self.ychenik)


        self.output_1()
        self.output_2()

    # вывод бд на экран
    def output_1(self):
        self.model = QSqlTableModel(self)
        self.model.setTable("ученики")
        self.model.select()
        self.ui_prepodavatel.tableView_prep.setModel(self.model)

    def output_2(self):
        self.model = QSqlTableModel(self)
        self.model.setTable("ученики")
        self.model.select()
        self.ui_ychenik.tableView.setModel(self.model)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    x = Main()
    x.avtorizacia.show()
    sys.exit(app.exec())
